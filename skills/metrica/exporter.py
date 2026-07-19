"""
Сохранение результатов в многостраничный Excel с форматированием.
"""

import pandas as pd
from openpyxl.styles import PatternFill


def to_excel(data_sheets, output_path):
    """Сохраняет данные в многостраничный Excel с цветовой индикацией."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows in data_sheets.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            # Цветовая индикация отказов > 70% (на любом листе с bounce_rate)
            if "bounce_rate" in df.columns:
                red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                bounce_col = None
                for idx, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
                    if col[0].value == "bounce_rate":
                        bounce_col = idx
                        break

                if bounce_col:
                    for row in worksheet.iter_rows(min_row=2, min_col=bounce_col, max_col=bounce_col):
                        cell = row[0]
                        try:
                            if float(cell.value) > 70:   # отказы в процентах
                                cell.fill = red
                        except (ValueError, TypeError):
                            pass

            # Дашборд «Сводка»: рост зелёным, падение красным (для отказов — наоборот)
            if sheet_name == "Сводка" and "Δ %" in df.columns:
                green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red2 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                delta_col = list(df.columns).index("Δ %") + 1
                for r in range(len(df)):
                    val = df.iloc[r]["Δ %"]
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        continue
                    good = val > 0
                    if "Отказ" in str(df.iloc[r]["Метрика"]):   # рост отказов — плохо
                        good = val < 0
                    worksheet.cell(row=r + 2, column=delta_col).fill = green if good else red2
