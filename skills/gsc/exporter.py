"""
Сохранение отчётов в Excel: каждый отчёт — отдельная вкладка.
CTR подсвечивается цветом (зелёный > 3%, жёлтый 1-3%, красный < 1%).
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def to_excel(sheets, output_path):
    """
    sheets: {"имя вкладки": [ {колонка: значение, ...}, ... ], ...}
    Пустые отчёты тоже получают вкладку (с пометкой), чтобы состав файла был стабильным.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, rows in sheets.items():
            sheet_name = name[:31]  # лимит Excel на имя листа
            df = pd.DataFrame(rows) if rows else pd.DataFrame({"нет данных": []})
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], df)


def _format_sheet(ws, df):
    """Шапка, ширина колонок, заморозка первой строки, подсветка CTR."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER
    ws.freeze_panes = "A2"

    for idx, col in enumerate(df.columns, 1):
        sample = df[col].astype(str).head(100).map(len).max() if len(df) else 0
        width = min(60, max(10, int(max(sample or 0, len(str(col)))) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width

    if "ctr" in df.columns:
        ctr_col = list(df.columns).index("ctr") + 1
        for row in ws.iter_rows(min_row=2, min_col=ctr_col, max_col=ctr_col):
            cell = row[0]
            try:
                value = float(cell.value)
            except (ValueError, TypeError):
                continue
            cell.number_format = "0.0%"
            cell.fill = GREEN if value > 0.03 else YELLOW if value >= 0.01 else RED
