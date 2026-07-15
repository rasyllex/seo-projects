"""
Скилл: Съём позиций.
Точка входа. Оркестратор: один SERP на ключ → позиции + конкуренты.
"""

import argparse
import io
import sys
from datetime import date
from pathlib import Path

import pandas as pd

# На Windows консоль по умолчанию cp1252 — ломает кириллицу. Переключаем на UTF-8.
if sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from fetcher import (
    fetch_serps, positions_from_serps, competitors_from_serps, _normalize_domain,
    snippets_from_serps, ngrams_from_serps, title_freq_from_serps, features_from_serps,
    highlights_from_serps, url_champions_from_serps, clusters_from_serps,
    url_champions_by_cluster,
)
from cache import load_cache, save_cache
from validate_input import validate_excel
from region_resolver import resolve_yandex, resolve_google, xmlriver_country_id


def _pick(matches, city, label):
    """Единственное совпадение региона или None (с выводом причины)."""
    if not matches:
        print(f"[ERROR] Город/регион '{city}' не найден в справочнике {label}.")
        return None
    if len(matches) > 1:
        print(f"[ERROR] Несколько совпадений для '{city}' в {label}:")
        for m in matches:
            print(f"  {m[0]}: {m[1]}")
        print("Уточните название или используйте --region <код>.")
        return None
    return matches[0]


def _load_keywords(path):
    """Читает ключи из .csv или .xlsx (баг: раньше всегда read_excel → падал на csv)."""
    if Path(path).suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    return df["keyword"].dropna().astype(str).tolist()


def _style_excel(path):
    """Оформление отчёта:
    - во всех листах закрепляется первая строка (заголовки);
    - строка, с которой начинается новый кластер, заливается светло-зелёным
      (в листах, где есть колонка cluster / cluster_id)."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    fill = PatternFill("solid", fgColor="E2EFDA")   # светло-зелёный Excel
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"                      # шапка всегда на экране
        header = [c.value for c in ws[1]]
        col = next((i for i, h in enumerate(header, 1)
                    if h in ("cluster", "cluster_id")), None)
        if col is None:
            continue
        prev = object()                             # сентинел: первая строка данных тоже граница
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val != prev:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = fill
            prev = val
    wb.save(path)


def _print_stats(results):
    total = len(results)
    errors = sum(1 for r in results if r["position"] == -1)
    found = sum(1 for r in results if r["position"] > 0)
    top10 = sum(1 for r in results if 1 <= r["position"] <= 10)
    top100 = sum(1 for r in results if 1 <= r["position"] <= 100)
    avg = sum(r["position"] for r in results if r["position"] > 0) / found if found else 0
    print(f"Всего ключей: {total}")
    if errors:
        print(f"Ошибок запроса (position=-1, не кэшированы): {errors}")
    print(f"Найдено в выдаче (ТОП-100): {found}")
    print(f"В ТОП-10: {top10}  |  В ТОП-100: {top100}  |  Средняя позиция: {avg:.1f}")


def main():
    parser = argparse.ArgumentParser(description="Съём позиций через XMLRiver")
    parser.add_argument("--input", required=True, help="Excel/CSV с колонкой keyword")
    parser.add_argument("--domain", required=True, help="Домен для проверки")
    parser.add_argument("--region", default=213, type=int, help="Код региона Яндекса (213 = Москва)")
    parser.add_argument("--city", help="Название города/региона. Альтернатива --region.")
    parser.add_argument("--engine", default="yandex", choices=["yandex", "google", "both"],
                        help="Поисковая система (both — обе)")
    parser.add_argument("--real", action="store_true", help="Реальный API (иначе мок)")
    parser.add_argument("--output", default=None,
                        help="Путь к Excel. По умолчанию: data/<домен>_<ДД.ММ.ГГГГ>.xlsx")
    parser.add_argument("--threads", default=8, type=int, help="Потоки (по умолчанию 8)")
    parser.add_argument("--competitors", action=argparse.BooleanOptionalAction, default=True,
                        help="Собирать ТОП-N конкурентов из того же SERP (по умолчанию ДА; --no-competitors — отключить)")
    parser.add_argument("--top-n", default=10, type=int, help="Сколько конкурентов (по умолчанию 10)")
    parser.add_argument("--text", action=argparse.BooleanOptionalAction, default=True,
                        help="Сниппеты, n-граммы, леммы title, Google-фичи из того же SERP (по умолчанию ДА)")
    parser.add_argument("--cluster-method", default="middle", choices=["middle", "hard"],
                        help="Кластеризация по ТОПам: middle (по умолчанию) или hard (клика)")
    parser.add_argument("--cluster-threshold", default=4, type=int,
                        help="Минимум общих URL для связи запросов (по умолчанию 4)")
    parser.add_argument("--cluster-depth", default=20, type=int,
                        help="Срез ТОПа для сравнения URL (по умолчанию 20)")
    args = parser.parse_args()

    # 1. Валидация и загрузка ключей
    print("-> Валидация входных данных...")
    ok, err = validate_excel(args.input)
    if not ok:
        print(f"[ERROR] {err}")
        return
    keywords = _load_keywords(args.input)
    print(f"[OK] Ключей для проверки: {len(keywords)}")

    engines = ["yandex", "google"] if args.engine == "both" else [args.engine]

    # 2. Разрешение региона
    region_yandex = args.region
    google_loc = None
    google_country = None
    if args.city:
        if "yandex" in engines:
            m = _pick(resolve_yandex(args.city), args.city, "yandex")
            if m is None:
                return
            region_yandex, name = m[0], m[1]
            print(f"-> Регион Яндекса: {name} (id={region_yandex})")
        if "google" in engines:
            g = resolve_google(args.city)
            if len(g) == 1:
                google_loc, gname, iso = g[0]
                google_country = xmlriver_country_id(iso)
                print(f"-> Google: {gname} (loc={google_loc}, country={google_country}/{iso})")
            elif len(g) > 1:
                _pick(g, args.city, "google")   # покажет варианты
                return
            else:
                print(f"[WARN] Google: город '{args.city}' не разрешён, беру дефолтный регион.")

    # 3. Съём: один SERP на ключ → позиции, конкуренты, тексты, фичи
    cache = load_cache()
    by_engine = {}
    serps_by_engine = {}
    for engine in engines:
        print(f"-> {engine}: домен {args.domain}")
        serps = fetch_serps(
            keywords, engine=engine, region=region_yandex,
            google_loc=google_loc, google_country=google_country,
            use_real=args.real, cache=cache, max_workers=args.threads,
            mock_domain=None if args.real else _normalize_domain(args.domain),
        )
        serps_by_engine[engine] = serps
        d = {"positions": positions_from_serps(serps, args.domain)}
        if args.competitors:
            d["competitors"] = competitors_from_serps(serps, args.top_n)
        # кластеризация по ТОПам (на кэшированных SERP, без запросов);
        # url нашего сайта по каждому ключу — из уже посчитанных позиций
        url_map = {r["keyword"]: r["url"] for r in d["positions"]}
        clusters, kw2cl, cl_rows = clusters_from_serps(
            serps, method=args.cluster_method,
            threshold=args.cluster_threshold, depth=args.cluster_depth, url_map=url_map)
        d["clusters"] = cl_rows
        d["url_champions"] = url_champions_by_cluster(serps, clusters)
        print(f"   кластеров ({args.cluster_method}, порог {args.cluster_threshold}, "
              f"ТОП-{args.cluster_depth}): {len(clusters)}")
        if args.text:
            # все текстовые агрегаты — ВНУТРИ кластера, не по всему ядру
            d["snippets"] = snippets_from_serps(serps, kw2cl)
            d["ngrams"] = ngrams_from_serps(serps, clusters)
            d["titles"] = title_freq_from_serps(serps, clusters)
            d["highlights"] = highlights_from_serps(serps, clusters)
            feats = features_from_serps(serps)
            if feats:
                d["features"] = feats
        by_engine[engine] = d
    save_cache(cache)

    # 3b. Сводный лист по всем ПС сразу
    champions = url_champions_from_serps(serps_by_engine)          # url + частота в ТОПе

    # 4. Имя файла: домен + дата
    if args.output:
        output_path = Path(args.output)
    else:
        slug = _normalize_domain(args.domain)
        output_path = Path(__file__).parents[2] / "data" / "positions" / f"{slug}_{date.today().strftime('%d.%m.%Y')}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 5. Запись листов
    if len(engines) == 1 and not args.competitors and not args.text:
        pd.DataFrame(by_engine[engines[0]]["positions"]).to_excel(output_path, index=False)
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            def dump(key, prefix):
                for engine in engines:
                    d = by_engine[engine]
                    if key in d:
                        pd.DataFrame(d[key]).to_excel(writer, sheet_name=f"{prefix}_{engine}"[:31], index=False)
            dump("positions", "positions")
            dump("clusters", "clusters")
            dump("competitors", "competitors")
            dump("snippets", "snippets")
            dump("ngrams", "ngrams")
            dump("titles", "titles")
            dump("highlights", "highlights")
            dump("features", "features")
            dump("url_champions", "url_champions")   # чемпионы внутри кластеров
            # сводный лист (по всем ПС сразу, без кластеров)
            if champions:
                pd.DataFrame(champions).to_excel(writer, sheet_name="url_champions", index=False)
        _style_excel(output_path)   # закреплённая шапка + зелёные границы кластеров
    print(f"[OK] Результат сохранён: {output_path}")

    # 6. Статистика
    for engine in engines:
        print(f"\n=== Статистика {engine} ===")
        _print_stats(by_engine[engine]["positions"])


if __name__ == "__main__":
    main()
