"""
Скилл: Проверка индексации
Точка входа.
"""

import argparse
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill

from checker import check_urls
from cache import load_cache, save_cache


def main():
    parser = argparse.ArgumentParser(description="Проверка индексации страниц")
    parser.add_argument("--input", required=True, help="Excel с колонкой url")
    parser.add_argument("--engines", default="google,yandex", help="Поисковики через запятую")
    parser.add_argument("--output", default=None, help="Путь к выходному Excel")
    parser.add_argument("--real", action="store_true", help="Использовать реальный API")
    args = parser.parse_args()

    engines = [e.strip() for e in args.engines.split(",")]

    # Чтение URL (поддержка CSV и Excel)
    input_path = Path(args.input)
    if input_path.suffix.lower() == '.csv':
        df = pd.read_csv(args.input)
    else:
        df = pd.read_excel(args.input)
    
    urls = df["url"].dropna().astype(str).tolist()
    print(f"-> URL для проверки: {len(urls)}")

    # Загрузка кэша
    cache = load_cache()

    # Проверка
    results = check_urls(urls, engines, use_real=args.real, cache=cache)

    # Сохранение кэша
    save_cache(cache)

    # Определение пути выходного файла
    if args.output:
        output_path = Path(args.output)
    else:
        # Дефолтный путь: data/index-check/output.xlsx
        output_path = Path(__file__).parents[2] / "data" / "index-check" / "output.xlsx"
    
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df_result = pd.DataFrame(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False)
        worksheet = writer.sheets["Sheet1"]

        # Цветовая индикация
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=1 + len(engines)):
            values = [cell.value for cell in row]
            true_count = sum(1 for v in values if v is True)

            if true_count == len(engines):
                fill = green
            elif true_count == 0:
                fill = red
            else:
                fill = yellow

            for cell in row:
                cell.fill = fill

    print(f"[OK] Результат сохранён: {output_path}")

    # Статистика
    total = len(results)
    for engine in engines:
        indexed = sum(1 for r in results if r.get(engine) is True)
        print(f"{engine}: проиндексировано {indexed} / {total}")


if __name__ == "__main__":
    main()
